

# -*- coding: UTF-8 -*-  
#为什么这个UTF-8代码不能放在第一行？

from rasterstats import point_query
import geopandas as gpd
from collections import Counter
import openpyxl
#请注意，Geopandas、Rasterstats,fiona,GDAL等Python用于数据处理的包的安装过程需要非常多的包依赖
#并且不能使用Pip install 命令直接安装，需要事先下载whl后，按照依赖顺序逐步安装，详细的环境配置请Google
#或许可以看一下conda安装怎么安
#openpyxl库用于提取出的List文件写入xlxs，请注意，Rasterstats包提取的数据格式为List（列表格式）

import sys
import json
import os
import pandas as pd
import csv
#本代码用于快速的以ArcGIS矢量点文件提取对应栅格的栅格值，并且将栅格值加入到指定的XLXS列中。

pointData = gpd.read_file(r'G:\Bachelor_Degree_Thesis\DATA\DATA_Points\DATA.shp')
point_raster_value = point_query(pointData['geometry'],r'G:\Bachelor_Degree_Thesis\DATA\PSM\tem_1985.tif',interpolate='nearest')
#rasterstats命令,该包是本提取方法的核心，不使用ArcPy，使用本包进行提取，本包需要的依赖包环境较为复杂，请注意。
#在此输入shp矢量点文件和栅格的文件路径
#本代码似乎只能提取基于地理坐标系的栅格，所以请注意栅格坐标系的转换！！
#2022.05.18,我的电脑炸了，我好痛苦，软件环境全崩溃了，这到底是什么人间疾苦，还有天理吗还有王法吗要死了真是，靠！！
#而且A-SOUL也出事了，为什么从西藏下来后就这么倒霉倒霉倒霉倒霉倒霉倒霉倒霉倒霉倒霉倒霉倒霉倒霉倒霉倒霉倒霉倒霉倒霉倒霉倒霉倒霉倒霉倒霉倒霉倒霉倒霉倒霉所有的事情都不顺利我要死了
#这日子真的没法过了

# print(point_raster_value)
# 显示

counter = Counter(point_raster_value)
print(counter)
#显示栅格频率，使用counter函数

#加载文件，修改xlsx表格
wb = openpyxl.load_workbook("C:/Users/Charlie林川/Desktop/PSM.xlsx")

#获得sheet名称
sheetNames = wb.sheetnames
print(sheetNames)

#sheetName1 = sheetNames[0]
#根据名称获取第一个sheet
#sheet1 = wb[sheetName1]
#根据索引获得第一个sheet
sheet1 = wb.worksheets[0]


#excel中单元格为B2开始，即第2列，第2行
#源代码来自：https://wenwen.soso.com/z/q912484938.htm
# i+X，Y中，X是第X行，Y是第Y列
for i in range(len(point_raster_value)):
    sheet1.cell(i+2, 7).value = point_raster_value[i]

#保存数据，如果提示权限错误，需要关闭打开的excel
wb.save("C:/Users/Charlie林川/Desktop/PSM.xlsx")
